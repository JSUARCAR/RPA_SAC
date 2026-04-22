"""
Módulo de manejo de datos para el sistema SAC Automation.

Implementa la clase DataHandler para leer credenciales del archivo Excel,
iterar sobre registros de tareas pendientes y actualizar resultados.

Classes
-------
DataHandler
    Gestiona la lectura y escritura de datos en el archivo Excel de control.

Functions
---------
Ninguna públicas.

Examples
--------
>>> from src.data_handler import DataHandler
>>> handler = DataHandler(config)
>>> for record in handler.iterate_records():
...     print(record)
"""

import logging
from openpyxl import load_workbook
from datetime import datetime
import shutil
import os


class DataHandler:
    """
    Gestiona la lectura y escritura de datos en el archivo Excel de control.

    Proporciona metodos para cargar el archivo Excel, leer tareas pendientes,
    actualizar estados y guardar cambios de forma segura.

    Examples
    --------
    >>> handler = DataHandler(config)
    >>> for task in handler.iterate_records():
    ...     print(task)
    """

    def __init__(self, initial_config):
        """
        Inicializa el manejador de datos.

        Copia el archivo Excel fuente, carga el workbook y extrae
        la configuracion de la hoja PARAMETROS_LOCALES.

        Parameters
        ----------
        initial_config : configparser.SectionProxy
            Configuracion inicial desde config.ini.
        """
        self.initial_config = initial_config
        self.excel_path = self._copy_source_excel()
        self.workbook = load_workbook(self.excel_path)
        self.ws_control = self.workbook["CONTROL_DIARIO_PQR"]
        self.ws_config = self.workbook["PARAMETROS_LOCALES"]
        self.ws_log = self.workbook["LOG"]
        self.config = self._load_config_from_excel()
        logging.info(
            f"Manejador de datos inicializado con el archivo: {self.excel_path}"
        )

    def _copy_source_excel(self):
        """
        Copia el archivo Excel fuente a un directorio de trabajo.

        Crea un backup del archivo con timestamp para mantener
        trazabilidad de las ejecuciones.

        Returns
        -------
        str
            Ruta completa del archivo copiado.
        """
        source_path = self.initial_config.get("source_excel_path")
        dest_dir = self.initial_config.get("dest_dir")

        if not source_path or not dest_dir:
            raise ValueError(
                "Las rutas 'source_excel_path' y 'dest_dir' no pueden estar vacías en config.ini"
            )

        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Archivo fuente no encontrado: {source_path}")

        os.makedirs(dest_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest_filename = f"Control_PQR_JSUARCAR_EMTELCO_{timestamp}.xlsx"
        dest_path = os.path.join(dest_dir, dest_filename)

        shutil.copy2(source_path, dest_path)
        logging.info(f"Archivo copiado de forma segura a: {dest_path}")
        return dest_path

    def _load_config_from_excel(self):
        """Carga la configuración específica desde la hoja PARAMETROS_LOCALES."""
        try:
            return {
                "anexos_mercurio_path": self.ws_config["K3"].value,
                "anexos_correo_path": self.ws_config["K2"].value,
                "sac_url": "https://sac.corp.epm.com.co/GEN/Vistas/Login/LOGIN_GEN.aspx",
                "sac_admin_url": "https://sac.corp.epm.com.co/SAC/Vistas/App/PRO_ADMPRO.aspx",
                "user": self.ws_config["M2"].value,
                "password": self.ws_config["N2"].value,
            }
        except KeyError:
            raise ValueError(
                "La hoja 'PARAMETROS_LOCALES' o sus celdas de configuración no se encontraron."
            )

    def read_credentials(self):
        """
        Lee las credenciales del archivo Excel de configuracion.

        Obtiene usuario y contrasena de la configuracion cargada
        desde la hoja PARAMETROS_LOCALES.

        Returns
        -------
        dict
            Diccionario con claves 'user' y 'password'.
        """
        # Nota: Este método no expone las credenciales directamente,
        # solo confirma que están disponibles para uso interno.
        user = self.config.get("user")
        password = self.config.get("password")
        if not user or not password:
            raise ValueError("Credenciales no encontradas en la configuración.")
        logging.info("Credenciales leídas exitosamente (no expuestas).")
        return {
            "user": user,
            "password": password,
        }  # En producción, considerar no retornar la contraseña

    def iterate_records(self):
        """
        Itera sobre los registros de la base de datos de solicitudes.

        Yields
        ------
        dict
            Diccionario representando un registro de solicitud
            con los datos de la tarea.
        """
        tasks = self.get_pending_tasks()
        for task in tasks:
            yield task

    def update_results(self, row_index, status, message=""):
        """
        Actualiza los resultados en el archivo Excel de trazabilidad.

        Actualiza el estado de la tarea y guarda los cambios
        automaticamente en el archivo Excel.

        Parameters
        ----------
        row_index : int
            Indice de la fila a actualizar en la hoja de control.
        status : str
            Nuevo estado: 'ANEXO RPA', 'SIN ANEXO RPA', o 'ERROR'.
        message : str, optional
            Mensaje descriptivo para casos de error.
        """
        self.update_task_status(row_index, status, message)
        self.save()

    def get_pending_tasks(self):
        """
        Lee la hoja de control y retorna las tareas pendientes.

        Filtra las tareas de la fecha actual con medios validos
        y estados que requieren procesamiento.

        Returns
        -------
        list
            Lista de diccionarios con las tareas pendientes.
        """
        tasks = []
        fecha_hoy = datetime.now().date()

        # La tabla TB_CONTROL_DIARIO inicia desde la fila 9
        for row_idx, row in enumerate(self.ws_control.iter_rows(min_row=9), start=9):
            # Extraer valores de las celdas
            fecha_fila_raw = row[5].value
            medio = str(row[6].value).strip() if row[6].value else ""
            num_proceso = str(row[8].value).strip() if row[8].value else ""
            anexo_estado = str(row[31].value).strip() if row[31].value else ""

            # --- Lógica de validación de fecha ---
            fecha_fila = None
            if isinstance(fecha_fila_raw, datetime):
                fecha_fila = fecha_fila_raw.date()
            elif isinstance(fecha_fila_raw, str):
                for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                    try:
                        fecha_fila = datetime.strptime(
                            fecha_fila_raw.strip(), fmt
                        ).date()
                        break
                    except (ValueError, TypeError):
                        continue

            # --- Criterios de filtrado ---
            if (
                fecha_fila
                and fecha_fila == fecha_hoy
                and medio in ["Escrito", "M - E-Mail"]
                and num_proceso
                and num_proceso not in ["", ".", "0"]
                and anexo_estado in ["", "SIN ANEXO RPA", "ERROR"]
            ):
                task_data = {
                    "row_index": row_idx,
                    "asunto": row[3].value,
                    "medio": medio,
                    "no_pqr": row[7].value,
                    "proceso_sac": num_proceso,
                }
                tasks.append(task_data)
                logging.debug(f"Fila {row_idx} añadida a las tareas pendientes.")

        logging.info(f"Se encontraron {len(tasks)} tareas pendientes.")
        return tasks

    def update_task_status(self, row_index, status, message=""):
        """
        Actualiza el estado de una tarea en la hoja de control.

        Escribe el estado en la columna AH y la fecha/hora actual
        en la columna AI de la hoja CONTROL_DIARIO_PQR.

        Parameters
        ----------
        row_index : int
            Indice de la fila a actualizar.
        status : str
            Nuevo estado: 'ANEXO RPA', 'SIN ANEXO RPA', o 'ERROR'.
        message : str, optional
            Mensaje descriptivo del resultado.
        """
        try:
            self.ws_control.cell(row=row_index, column=32).value = status
            self.ws_control.cell(
                row=row_index, column=33
            ).value = datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
            if status == "ERROR":
                # Podríamos usar otra columna para el mensaje de error si existiera
                pass
            logging.debug(f"Fila {row_index} actualizada con estado: {status}")
        except Exception as e:
            logging.error(
                f"No se pudo actualizar el estado de la fila {row_index}: {e}"
            )

    def save(self):
        """
        Guarda los cambios realizados en el archivo Excel.

        Persiste todos los cambios realizados al workbook
        en el archivo de control.

        Returns
        -------
        None

        Raises
        ------
        Exception
            Si ocurre un error al guardar el archivo.
        """
        try:
            self.workbook.save(self.excel_path)
            logging.info("Cambios en el archivo Excel guardados correctamente.")
        except Exception as e:
            logging.error(f"No se pudo guardar el archivo Excel: {e}")
            raise
