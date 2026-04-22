"""
Módulo de acceso a datos para el sistema SAC Automation.

Encapsula toda la lógica de interacción con el archivo Excel de control,
proveyendo una interfaz limpia para leer tareas pendientes y actualizar
resultados de procesamiento.

Classes
-------
DataHandler
    Gestiona la lectura y escritura de datos en el archivo Excel de control.
"""

import logging
from openpyxl import load_workbook
from datetime import datetime
import shutil
import os


class DataHandler:
    """
    Gestiona la lectura y escritura de datos en el archivo Excel.

    Proporciona métodos para leer tareas pendientes, actualizar estados
    y guardar cambios en el archivo de control.
    """

    def __init__(self, initial_config):
        """Inicializa el manejador de datos."""
        self.initial_config = initial_config
        self.excel_path = self._copy_source_excel()
        self.workbook = load_workbook(self.excel_path)
        self.ws_control = self.workbook["CONTROL_DIARIO_PQR"]
        self.ws_config = self._get_config_sheet()
        self.ws_log = self.workbook["LOG"]
        self.config = self._load_config_from_excel()
        logging.info(f"Manejador de datos inicializado con archivo de trabajo")

    def _copy_source_excel(self):
        """Copia el archivo Excel fuente a un directorio de trabajo."""
        logger = logging.getLogger("DataHandler")

        source_path = self.initial_config.get("source_excel_path")
        dest_dir = self.initial_config.get("dest_dir")

        if not source_path or not dest_dir:
            logger.critical("Configuración incompleta: rutas de Excel no definidas")
            raise ValueError(
                "source_excel_path y dest_dir son requeridos en config.ini"
            )

        if not os.path.exists(source_path):
            logger.critical(f"Archivo fuente no encontrado: [{source_path}]")
            raise FileNotFoundError(f"Archivo de control no encontrado: {source_path}")

        os.makedirs(dest_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest_filename = f"Control_PQR_JSUARCAR_EMTELCO_{timestamp}.xlsx"
        dest_path = os.path.join(dest_dir, dest_filename)

        shutil.copy2(source_path, dest_path)
        logger.info(f"✓ Archivo de trabajo preparado: [{dest_filename}]")
        return dest_path

    def _get_config_sheet(self):
        """Obtiene la hoja de configuración del workbook."""
        logger = logging.getLogger("DataHandler")

        if "PARAMETROS_LOCALES" in self.workbook.sheetnames:
            return self.workbook["PARAMETROS_LOCALES"]

        logger.debug("Buscando hoja de configuración por codename 'Hoja15'...")
        for ws in self.workbook.worksheets:
            if ws.sheet_properties.codeName == "Hoja15":
                logger.info(f"✓ Hoja de configuración localizada: [{ws.title}]")
                return ws

        logger.critical("Hoja de configuración no encontrada en workbook")
        raise ValueError("No se encontró PARAMETROS_LOCALES ni Hoja15")

    def _load_config_from_excel(self):
        """Carga la configuración específica desde PARAMETROS_LOCALES."""
        logger = logging.getLogger("DataHandler")

        try:
            logger.debug("Leyendo configuración de PARAMETROS_LOCALES...")

            user_raw = self.ws_config["M2"].value
            password_raw = self.ws_config["N2"].value
            anexos_mercurio = self.ws_config["K3"].value
            anexos_correo = self.ws_config["K2"].value

            if not user_raw or str(user_raw).strip() == "":
                logger.critical("Celda M2 (Usuario) está vacía en PARAMETROS_LOCALES")
                logger.critical("Acción: Ingresar credencial de acceso en M2")
                raise ValueError("Celda M2 vacía: Usuario requerido")

            if not password_raw or str(password_raw).strip() == "":
                logger.critical(
                    "Celda N2 (Contraseña) está vacía en PARAMETROS_LOCALES"
                )
                logger.critical("Acción: Ingresar contraseña en N2")
                raise ValueError("Celda N2 vacía: Contraseña requerida")

            user = str(user_raw).strip()
            password = str(password_raw).strip()

            if len(password) < 4:
                logger.warning(
                    f"Contraseña con longitud inusual: {len(password)} caracteres"
                )

            logger.debug("✓ Credenciales validadas correctamente")

            return {
                "anexos_mercurio_path": anexos_mercurio,
                "anexos_correo_path": anexos_correo,
                "sac_url": "https://sac.corp.epm.com.co/GEN/Vistas/Login/LOGIN_GEN.aspx",
                "sac_admin_url": "https://sac.corp.epm.com.co/SAC/Vistas/App/PRO_ADMPRO.aspx",
                "user": user,
                "password": password,
            }
        except KeyError as e:
            logger.critical(f"Celda de configuración no encontrada: {e}")
            raise ValueError(f"Configuración incompleta: {e}")

    def get_pending_tasks(self):
        """
        Lee la hoja de control y retorna una lista de diccionarios con tareas pendientes.

        Filtra las tareas basándose en:
        - Fecha igual a la fecha actual
        - Medio de recepción: 'Escrito' o 'M - E-Mail'
        - Estado actual: vacío, 'SIN ANEXO RPA', o 'ERROR'
        - Número de proceso válido
        """
        logger = logging.getLogger("DataHandler")
        tasks = []
        fecha_hoy = datetime.now().date()

        logger.debug(f"Consultando tareas con fecha: {fecha_hoy}")

        for row_idx, row in enumerate(self.ws_control.iter_rows(min_row=9), start=9):
            fecha_fila_raw = row[5].value
            medio = str(row[6].value).strip() if row[6].value else ""
            num_proceso = str(row[8].value).strip() if row[8].value else ""
            anexo_estado = str(row[31].value).strip() if row[31].value else ""

            fecha_fila = None
            if isinstance(fecha_fila_raw, datetime):
                fecha_fila = fecha_fila_raw.date()
            elif isinstance(fecha_fila_raw, str):
                for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                    try:
                        fecha_fila = datetime.strptime(
                            fecha_fila_raw.strip(), fmt
                        ).date()
                        break
                    except (ValueError, TypeError):
                        continue

            if (
                fecha_fila
                and fecha_fila == fecha_hoy
                and medio in ["Escrito", "M - E-Mail"]
                and num_proceso
                and num_proceso not in ["", ".", "0"]
                and anexo_estado in ["", "SIN ANEXO RPA", "ERROR"]
            ):
                task_data = {
                    "row_index": row_idx,
                    "asunto": row[3].value,
                    "medio": medio,
                    "no_pqr": row[7].value,
                    "proceso_sac": num_proceso,
                }
                tasks.append(task_data)
                logger.debug(f"Fila {row_idx} añadida a cola de trabajo")

        logger.info(
            f"✓ {len(tasks)} tarea(s) pendiente(s) encontrada(s) para fecha {fecha_hoy}"
        )
        return tasks

    def update_task_status(self, row_index, status, message=""):
        """Actualiza el estado de una tarea en la hoja de control."""
        logger = logging.getLogger("DataHandler")

        try:
            self.ws_control.cell(row=row_index, column=32).value = status
            self.ws_control.cell(
                row=row_index, column=33
            ).value = datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")

            if status == "ERROR":
                logger.warning(
                    f"Estado ERROR registrado | Fila: {row_index} | Mensaje: {message[:50]}..."
                )
            else:
                logger.debug(
                    f"Estado actualizado | Fila: {row_index} | Nuevo estado: {status}"
                )

        except Exception as e:
            logger.error(
                f"Fallo al actualizar estado | Fila: {row_index} | Error: {type(e).__name__}"
            )

    def save(self):
        """Guarda los cambios en el archivo Excel."""
        logger = logging.getLogger("DataHandler")

        try:
            self.workbook.save(self.excel_path)
            logger.info("✓ Cambios guardados en archivo Excel")
        except Exception as e:
            logger.error(f"Fallo al guardar archivo Excel: {type(e).__name__}")
            raise
